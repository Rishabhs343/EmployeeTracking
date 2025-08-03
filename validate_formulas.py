#!/usr/bin/env python3
"""
Validate the formulas in the generated Excel file
"""

import openpyxl

def validate_excel_formulas():
    """Validate that all formulas in the Excel file are working correctly"""
    
    print("🔍 Validating Excel formulas...")
    
    # Load the generated workbook
    filename = "Employee_Performance_Tracker_Aug2025.xlsx"
    wb = openpyxl.load_workbook(filename, data_only=True)
    
    # Check Template sheet (clean data)
    print("\n📋 Validating Template sheet...")
    ws = wb["Template_Performance"]
    
    # Check a sample row (row 2)
    print("Row 2 values:")
    date_val = ws['A2'].value
    day_val = ws['B2'].value  
    meeting_hrs = ws['C2'].value
    assigned_hrs = ws['D2'].value
    completed_hrs = ws['E2'].value
    complexity = ws['F2'].value
    qa_factor = ws['G2'].value
    task_failed = ws['H2'].value
    leave_taken = ws['I2'].value
    available_hrs = ws['J2'].value
    efficiency = ws['K2'].value
    raw_points = ws['L2'].value
    ot_points = ws['M2'].value
    approved_points = ws['N2'].value
    
    print(f"  Date: {date_val}")
    print(f"  Day: {day_val}")
    print(f"  Meeting Hrs: {meeting_hrs}")
    print(f"  Assigned Hrs: {assigned_hrs}")
    print(f"  Completed Hrs: {completed_hrs}")
    print(f"  Complexity: {complexity}")
    print(f"  QA Factor: {qa_factor}")
    print(f"  Task Failed: {task_failed}")
    print(f"  Leave Taken: {leave_taken}")
    print(f"  Available Hrs: {available_hrs} (formula result)")
    print(f"  Efficiency: {efficiency:.1%} (formula result)")
    print(f"  Raw Points: {raw_points} (formula result)")
    print(f"  OT Points: {ot_points} (formula result)")
    print(f"  Approved Points: {approved_points} (formula result)")
    
    # Verify calculations manually
    expected_available = 9 - meeting_hrs  # 9 - 0 = 9
    expected_efficiency = completed_hrs / expected_available if expected_available > 0 else 0  # 9/9 = 1.0
    expected_raw = completed_hrs * complexity * qa_factor  # 9 * 1 * 1 = 9
    expected_ot = max(0, completed_hrs - assigned_hrs)  # max(0, 9-9) = 0
    expected_approved = expected_efficiency * expected_raw if task_failed != "Y" else 0  # 1.0 * 9 = 9
    
    print(f"\nManual verification:")
    print(f"  Expected Available: {expected_available} ✓" if available_hrs == expected_available else f"  Expected Available: {expected_available} ❌ Got: {available_hrs}")
    print(f"  Expected Efficiency: {expected_efficiency:.1%} ✓" if abs(efficiency - expected_efficiency) < 0.01 else f"  Expected Efficiency: {expected_efficiency:.1%} ❌ Got: {efficiency:.1%}")
    print(f"  Expected Raw Points: {expected_raw} ✓" if raw_points == expected_raw else f"  Expected Raw Points: {expected_raw} ❌ Got: {raw_points}")
    print(f"  Expected OT Points: {expected_ot} ✓" if ot_points == expected_ot else f"  Expected OT Points: {expected_ot} ❌ Got: {ot_points}")
    print(f"  Expected Approved: {expected_approved} ✓" if approved_points == expected_approved else f"  Expected Approved: {expected_approved} ❌ Got: {approved_points}")
    
    # Check summary section
    print("\n📊 Validating Summary section...")
    
    # Find summary section (around row 30)
    for row in range(25, 35):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and "Total Earned Points" in str(cell_value):
            summary_start = row
            break
    else:
        print("❌ Could not find summary section")
        return
    
    total_points = ws.cell(row=summary_start, column=3).value
    workdays = ws.cell(row=summary_start + 1, column=3).value
    base_salary = ws.cell(row=summary_start + 2, column=3).value
    bonus_rate = ws.cell(row=summary_start + 3, column=3).value
    monthly_bonus = ws.cell(row=summary_start + 4, column=3).value
    total_pay = ws.cell(row=summary_start + 5, column=3).value
    
    print(f"  Total Earned Points: {total_points}")
    print(f"  Workdays: {workdays}")
    print(f"  Base Salary: ₹{base_salary:,.2f}")
    print(f"  Bonus Rate per Point: ₹{bonus_rate:.2f}")
    print(f"  Monthly Bonus: ₹{monthly_bonus:,.2f}")
    print(f"  Total Pay: ₹{total_pay:,.2f}")
    
    # Manual verification of summary
    expected_total_points = 26 * 9  # 26 days * 9 points = 234
    expected_bonus_rate = (base_salary * 0.5) / (workdays * 10)  # (50000 * 0.5) / (26 * 10) = 96.15
    expected_bonus = min(total_points * expected_bonus_rate, base_salary * 0.5)
    expected_total_pay = base_salary + expected_bonus
    
    print(f"\nSummary verification:")
    print(f"  Expected Total Points: {expected_total_points} ✓" if total_points == expected_total_points else f"  Expected Total Points: {expected_total_points} ❌ Got: {total_points}")
    print(f"  Expected Bonus Rate: ₹{expected_bonus_rate:.2f} ✓" if abs(bonus_rate - expected_bonus_rate) < 0.01 else f"  Expected Bonus Rate: ₹{expected_bonus_rate:.2f} ❌ Got: ₹{bonus_rate:.2f}")
    print(f"  Expected Bonus: ₹{expected_bonus:,.2f} ✓" if abs(monthly_bonus - expected_bonus) < 1 else f"  Expected Bonus: ₹{expected_bonus:,.2f} ❌ Got: ₹{monthly_bonus:,.2f}")
    print(f"  Expected Total Pay: ₹{expected_total_pay:,.2f} ✓" if abs(total_pay - expected_total_pay) < 1 else f"  Expected Total Pay: ₹{expected_total_pay:,.2f} ❌ Got: ₹{total_pay:,.2f}")
    
    wb.close()
    
    print(f"\n✅ Validation completed for {filename}")
    print("\n📋 Expected results structure:")
    print("Date          Day  Meeting  Assigned  Completed  Complexity  QA  Failed  Leave  Available  Efficiency  Raw  OT  Approved")
    print("2025-08-01    Fri      0         9          9           1     1      N      N          9       100%      9   0         9")
    print("...")
    print("Summary:")
    print(f"Total Earned Points: {expected_total_points}")
    print(f"Workdays: 26")
    print(f"Base Salary: ₹50,000")
    print(f"Bonus Rate/Point: ₹{expected_bonus_rate:.2f}")
    print(f"Monthly Bonus: ₹{expected_bonus:,.2f}")
    print(f"Total Pay: ₹{expected_total_pay:,.2f}")

if __name__ == "__main__":
    validate_excel_formulas()