#!/usr/bin/env python3
"""
Employee Performance Tracker with Bonus Automation
Generates formula-protected Excel workbook for employee performance tracking
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.protection import WorkbookProtection
from datetime import datetime, timedelta
import calendar
import random


class PerformanceTracker:
    def __init__(self):
        self.workbook = Workbook()
        self.password = "secure123"
        self.base_salary = 50000
        
        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
    
    def get_workdays_for_month(self, year, month):
        """Get all Monday-Saturday dates for a given month"""
        workdays = []
        # Get the first day of the month
        first_day = datetime(year, month, 1)
        # Get the last day of the month
        last_day = datetime(year, month, calendar.monthrange(year, month)[1])
        
        current_date = first_day
        while current_date <= last_day:
            # Monday = 0, Sunday = 6, so exclude Sunday (6)
            if current_date.weekday() < 6:  # Monday to Saturday
                workdays.append(current_date)
            current_date += timedelta(days=1)
        
        return workdays
    
    def create_employee_sheet(self, employee_name, year=2025, month=8, sample_data=False):
        """Create a new sheet for an employee with performance tracking"""
        
        # Create worksheet
        ws = self.workbook.create_sheet(title=f"{employee_name}_Performance")
        
        # Get workdays for the month
        workdays = self.get_workdays_for_month(year, month)
        
        # Set up headers
        self.setup_headers(ws)
        
        # Set up daily data rows
        self.setup_daily_data(ws, workdays, sample_data)
        
        # Set up summary section
        self.setup_summary_section(ws, len(workdays))
        
        # Apply formatting
        self.apply_formatting(ws, len(workdays))
        
        # Set up data validation
        self.setup_data_validation(ws, len(workdays))
        
        # Apply conditional formatting
        self.apply_conditional_formatting(ws, len(workdays))
        
        # Set up protection
        self.setup_protection(ws, len(workdays))
        
        return ws
    
    def setup_headers(self, ws):
        """Set up column headers"""
        headers = [
            "Date", "Day", "Meeting Hrs", "Assigned Hrs", "Completed Hrs",
            "Complexity Factor", "QA Factor", "Task Failed (Y/N)", 
            "Leave Taken (Y/N)", "Available Hrs", "% Efficiency", 
            "Raw Points", "OT Points", "Approved Points"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def setup_daily_data(self, ws, workdays, sample_data=False):
        """Set up daily data rows with formulas"""
        
        for row_idx, date in enumerate(workdays, 2):
            row = row_idx
            
            # Column A: Date
            ws.cell(row=row, column=1, value=date.date())
            
            # Column B: Day (formula)
            ws.cell(row=row, column=2, value=f'=TEXT(A{row},"ddd")')
            
            # Columns C-I: Input fields with default values
            if sample_data:
                # Generate realistic sample data
                meeting_hrs = random.choice([0, 0.5, 1, 1.5, 2])
                assigned_hrs = 9
                completed_hrs = random.choice([7, 8, 9, 9.5, 10, 11])
                complexity = random.choice([0.8, 1.0, 1.2, 1.5])
                qa_factor = random.choice([0.9, 1.0, 1.1, 1.2])
                task_failed = random.choice(["N", "N", "N", "N", "Y"])  # 20% failure rate
                leave_taken = random.choice(["N", "N", "N", "N", "N", "N", "Y"])  # ~15% leave rate
                
                ws.cell(row=row, column=3, value=meeting_hrs)  # Meeting Hrs
                ws.cell(row=row, column=4, value=assigned_hrs)  # Assigned Hrs
                ws.cell(row=row, column=5, value=completed_hrs)  # Completed Hrs
                ws.cell(row=row, column=6, value=complexity)  # Complexity Factor
                ws.cell(row=row, column=7, value=qa_factor)  # QA Factor
                ws.cell(row=row, column=8, value=task_failed)  # Task Failed
                ws.cell(row=row, column=9, value=leave_taken)  # Leave Taken
            else:
                # Default values
                ws.cell(row=row, column=3, value=0)      # Meeting Hrs
                ws.cell(row=row, column=4, value=9)      # Assigned Hrs
                ws.cell(row=row, column=5, value=9)      # Completed Hrs
                ws.cell(row=row, column=6, value=1)      # Complexity Factor
                ws.cell(row=row, column=7, value=1)      # QA Factor
                ws.cell(row=row, column=8, value="N")    # Task Failed
                ws.cell(row=row, column=9, value="N")    # Leave Taken
            
            # Column J: Available Hrs (formula)
            ws.cell(row=row, column=10, value=f'=IF(I{row}="Y",0,9-C{row})')
            
            # Column K: % Efficiency (formula)
            ws.cell(row=row, column=11, value=f'=IF(J{row}=0,0,E{row}/J{row})')
            
            # Column L: Raw Points (formula)
            ws.cell(row=row, column=12, value=f'=E{row}*F{row}*G{row}')
            
            # Column M: OT Points (formula)
            ws.cell(row=row, column=13, value=f'=IF(E{row}>D{row},E{row}-D{row},0)')
            
            # Column N: Approved Points (formula)
            ws.cell(row=row, column=14, value=f'=IF(H{row}="Y",0,ROUND(K{row}*L{row},2))')
    
    def setup_summary_section(self, ws, num_workdays):
        """Set up summary calculation section"""
        summary_start_row = num_workdays + 4  # Start 2 rows after data
        
        # Calculate absolute cell references for summary formulas
        total_points_cell = f"C{summary_start_row}"        # Row 30: Total Earned Points
        workdays_cell = f"C{summary_start_row + 1}"        # Row 31: Total Workdays
        base_salary_cell = f"C{summary_start_row + 2}"     # Row 32: Base Salary
        bonus_rate_cell = f"C{summary_start_row + 3}"      # Row 33: Bonus Rate per Point
        monthly_bonus_cell = f"C{summary_start_row + 4}"   # Row 34: Monthly Bonus
        
        # Summary labels and formulas with correct cell references
        summary_data = [
            ("Total Earned Points", f"=SUM(N2:N{num_workdays + 1})"),
            ("Total Workdays", f"=COUNTA(A2:A{num_workdays + 1})"),
            ("Base Salary (₹)", self.base_salary),
            ("Bonus Rate per Point (₹)", f"={base_salary_cell}*0.5/({workdays_cell}*10)"),
            ("Monthly Bonus (₹)", f"=MIN({total_points_cell}*{bonus_rate_cell},{base_salary_cell}*0.5)"),
            ("Total Monthly Pay (₹)", f"={base_salary_cell}+{monthly_bonus_cell}")
        ]
        
        for i, (label, value) in enumerate(summary_data):
            row = summary_start_row + i
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            
            if isinstance(value, str) and value.startswith("="):
                ws.cell(row=row, column=3, value=value)
            else:
                ws.cell(row=row, column=3, value=value)
    
    def apply_formatting(self, ws, num_workdays):
        """Apply formatting to the worksheet"""
        
        # Set column widths
        column_widths = [12, 8, 10, 10, 10, 10, 10, 12, 12, 10, 10, 10, 10, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Format data table
        for row in range(2, num_workdays + 2):
            for col in range(1, 15):
                cell = ws.cell(row=row, column=col)
                
                # Alternating row colors
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                
                # Number formatting
                if col in [3, 4, 5]:  # Hours columns
                    cell.number_format = "0.0"
                elif col in [6, 7]:  # Factor columns
                    cell.number_format = "0.0"
                elif col == 11:  # Efficiency percentage
                    cell.number_format = "0.0%"
                elif col in [12, 13, 14]:  # Points columns
                    cell.number_format = "0.00"
        
        # Format summary section
        summary_start_row = num_workdays + 4
        for row in range(summary_start_row, summary_start_row + 6):
            ws.cell(row=row, column=3).number_format = "₹#,##0.00"
            
        # Add borders
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for row in range(1, num_workdays + 2):
            for col in range(1, 15):
                ws.cell(row=row, column=col).border = thin_border
    
    def setup_data_validation(self, ws, num_workdays):
        """Set up data validation for input fields"""
        
        # Validation rules
        validations = [
            (3, "decimal", "0,9", "Meeting hours must be between 0 and 9"),
            (4, "decimal", "1,12", "Assigned hours must be between 1 and 12"),
            (5, "decimal", "0,16", "Completed hours must be between 0 and 16"),
            (6, "decimal", "0.5,3.0", "Complexity factor must be between 0.5 and 3.0"),
            (7, "decimal", "0.3,2.0", "QA factor must be between 0.3 and 2.0"),
        ]
        
        for col, validation_type, formula_range, error_msg in validations:
            dv = DataValidation(type=validation_type, formula1=formula_range.split(',')[0], 
                              formula2=formula_range.split(',')[1])
            dv.error = error_msg
            dv.errorTitle = "Invalid Input"
            range_string = f"{get_column_letter(col)}2:{get_column_letter(col)}{num_workdays + 1}"
            dv.add(range_string)
            ws.add_data_validation(dv)
        
        # Dropdown validations for Y/N fields
        yn_validation = DataValidation(type="list", formula1='"Y,N"')
        yn_validation.error = "Please select Y or N"
        yn_validation.errorTitle = "Invalid Selection"
        
        # Task Failed column
        yn_validation.add(f"H2:H{num_workdays + 1}")
        ws.add_data_validation(yn_validation)
        
        # Leave Taken column
        yn_validation2 = DataValidation(type="list", formula1='"Y,N"')
        yn_validation2.error = "Please select Y or N"
        yn_validation2.errorTitle = "Invalid Selection"
        yn_validation2.add(f"I2:I{num_workdays + 1}")
        ws.add_data_validation(yn_validation2)
    
    def apply_conditional_formatting(self, ws, num_workdays):
        """Apply conditional formatting"""
        
        # Red highlighting for Task Failed = "Y"
        red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        task_failed_rule = CellIsRule(operator="equal", formula=['"Y"'], fill=red_fill)
        ws.conditional_formatting.add(f"H2:H{num_workdays + 1}", task_failed_rule)
        
        # Orange highlighting for Leave Taken = "Y"
        orange_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        leave_rule = CellIsRule(operator="equal", formula=['"Y"'], fill=orange_fill)
        ws.conditional_formatting.add(f"I2:I{num_workdays + 1}", leave_rule)
        
        # Green highlighting for efficiency > 100%
        green_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        efficiency_rule = CellIsRule(operator="greaterThan", formula=["1"], fill=green_fill)
        ws.conditional_formatting.add(f"K2:K{num_workdays + 1}", efficiency_rule)
    
    def setup_protection(self, ws, num_workdays):
        """Set up cell protection and sheet protection"""
        
        # First, lock all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.protection = openpyxl.styles.Protection(locked=True)
        
        # Unlock editable cells
        editable_columns = [3, 4, 5, 6, 7, 8, 9]  # C, D, E, F, G, H, I
        for col in editable_columns:
            for row in range(2, num_workdays + 2):
                ws.cell(row=row, column=col).protection = openpyxl.styles.Protection(locked=False)
        
        # Unlock Base Salary cell in summary
        summary_start_row = num_workdays + 4
        ws.cell(row=summary_start_row + 2, column=3).protection = openpyxl.styles.Protection(locked=False)
        
        # Protect the sheet
        ws.protection.sheet = True
        ws.protection.password = self.password
        ws.protection.formatCells = False
        ws.protection.formatColumns = False
        ws.protection.formatRows = False
        ws.protection.insertColumns = False
        ws.protection.insertRows = False
        ws.protection.deleteColumns = False
        ws.protection.deleteRows = False
    
    def add_employee(self, employee_name, year=2025, month=8, sample_data=False):
        """Add a new employee sheet to the workbook"""
        return self.create_employee_sheet(employee_name, year, month, sample_data)
    
    def save_workbook(self, filename="Employee_Performance_Tracker_Aug2025.xlsx"):
        """Save the workbook to file"""
        # Set workbook protection
        self.workbook.security = WorkbookProtection(workbookPassword=self.password, lockStructure=True)
        
        self.workbook.save(filename)
        print(f"✅ Workbook saved as: {filename}")
        return filename


def main():
    """Main function to generate the performance tracker"""
    
    print("🚀 Generating Employee Performance Tracker...")
    
    # Create tracker instance
    tracker = PerformanceTracker()
    
    # Add sample employee with sample data
    print("📊 Adding sample employee: John Doe")
    tracker.add_employee("John_Doe", year=2025, month=8, sample_data=True)
    
    # Add template employee without sample data
    print("📋 Adding template for new employees: Template")
    tracker.add_employee("Template", year=2025, month=8, sample_data=False)
    
    # Save the workbook
    filename = tracker.save_workbook()
    
    print("\n✅ Performance Tracker Generated Successfully!")
    print(f"📁 File: {filename}")
    print("🔒 Password: secure123")
    print("\n📋 Features included:")
    print("  ✓ 26 workdays pre-populated for August 2025")
    print("  ✓ All formulas for automatic calculations")
    print("  ✓ Sheet protection with password")
    print("  ✓ Data validation for input fields")
    print("  ✓ Conditional formatting for visual alerts")
    print("  ✓ Sample data for John Doe")
    print("  ✓ Clean template for new employees")
    print("\n🔧 To add new employees:")
    print("  1. Copy the 'Template' sheet")
    print("  2. Rename it with employee name")
    print("  3. Update Base Salary if needed")
    print("  4. Start entering daily data")
    
    return filename


def add_new_employee_programmatically(filename, employee_name, year=2025, month=8):
    """
    Function to add a new employee to existing workbook
    Usage example for extending the system
    """
    from openpyxl import load_workbook
    
    # Load existing workbook
    wb = load_workbook(filename)
    
    # Create new tracker instance and set the workbook
    tracker = PerformanceTracker()
    tracker.workbook = wb
    
    # Add new employee
    tracker.add_employee(employee_name, year, month, sample_data=False)
    
    # Save updated workbook
    tracker.save_workbook(filename)
    
    print(f"✅ Added new employee: {employee_name}")


if __name__ == "__main__":
    # Generate the initial workbook
    generated_file = main()
    
    # Example of adding another employee programmatically
    print("\n🔄 Example: Adding another employee...")
    add_new_employee_programmatically(generated_file, "Jane_Smith", 2025, 8)
    
    print("\n🎉 All done! Your Employee Performance Tracker is ready to use.")